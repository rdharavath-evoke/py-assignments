import openpyxl
import re
from openpyxl.styles.fonts import Font
import tabula

class readpdf:
    def method1(self):
        method=r"D:\py-assignments-new\py-assignments\Assigenmet3\tableform.pdf"
        data=tabula.read_pdf(method,pages=1)[0]
        print(data)
        data.to_excel(r"D:\py-assignments-new\py-assignments\Assigenmet3\pdfreader.xlsx",index=False)


class storedata(readpdf):
    
    
    def tabledata(self):
        wb=openpyxl.load_workbook(r"D:\py-assignments-new\py-assignments\Assigenmet3\pdfreader.xlsx")
        ws=wb.active
        ws['D1'].value="status"

        for i in range(2,40):
            numbers=ws.cell(row=i,column=3).value
            number=str(numbers)

            if re.compile(r'[01]\d{10}').search(number):
                ws['D{}'.format(i)]='Helpline_No'
                ws['D{}'.format(i)].font=Font(color="00FFFF")

            elif re.compile(r'\d{3,5}-\d{6,8}').search(number):
                ws['D{}'.format(i)]='Landline_No.'
                ws['D{}'.format(i)].font=Font(color="008000")
    
            elif re.compile(r'[36789]\d{9}').search(number):
                ws['D{}'.format(i)]='Mobile_No.'
                ws['D{}'.format(i)].font=Font(color="0000FF")

            elif re.compile(r'\d{3}').search(number):
                ws['D{}'.format(i)]='Emergency_Ambulance_No.'
                ws['D{}'.format(i)].font=Font(color="FF0000")

            else:
                ws['D{}'.format(i)]='None'
        #wb.save(r"D:\py-assignments-new\py-assignments\Assigenmet3\pdfreader2.xlsx")    

        for i in range(2,40):
            names=ws.cell(row=i,column=2).value
            name=str(names)

            if re.compile(r'^[M].*[h]$').search(name):
                ws['E{}'.format(i)]=name
                ws['E{}'.format(i)].font=Font(color="FF0000")
            else:
                ws['E{}'.format(i)]='None'  
                ws['E{}'.format(i)].font=Font(color="0000FF")      

        wb.save(r"D:\py-assignments-new\py-assignments\Assigenmet3\pdfreader2.xlsx")

sd=storedata()
print(sd.method1())
print(sd.tabledata())
import openpyxl
import logging
from openpyxl.styles.fonts import Font

path=r"D:\myfiles\new.xlsx"
wb=openpyxl.load_workbook(path)
ws=wb.active


logging.basicConfig(level=logging.DEBUG)

logging.debug("\n")

for c1,c2,c3,c4 in ws['A1':'D7']:
    logging.debug("    {0:7} {1:7} {2:7}  {3:7} ".format(c1.value,c2.value,c3.value, c4.value))
    

ws['E1'].value="Result"
#ws['E1'].font=Font(color="FF0000")


for i in range(2,8):
    
    ws['E{}'.format(i)].number_format='#,##0.00'

    result=ws.cell(row=i,column=4).value    
    if str(result).isdigit():
        ws['E{}'.format(i)]='=D{}/7'.format(i)
    else:
        ws['E{}'.format(i)]='error'
        ws['E{}'.format(i)].font=Font(color="FF0000")

wb.save('new.xlsx')
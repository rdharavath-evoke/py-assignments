
import re
import logging
import tabula
import openpyxl
from openpyxl.styles.fonts import Font

logging.basicConfig(level=logging.DEBUG)

wb=openpyxl.load_workbook(r"D:\py-assignments-new\py-assignments\assignment-2\pdfreader.xlsx")
ws=wb.active


file=r"D:\py-assignments-new\py-assignments\assignment-2\tableform.pdf"
data=tabula.read_pdf(file,pages=1)[0]
print(data)

data.to_excel(r"D:\py-assignments-new\py-assignments\assignment-2\pdfreader.xlsx",index=False)

ws['D1'].value="status"

for i in range(2,40):
    numbers=ws.cell(row=i,column=3).value
    number=str(numbers)

    if re.compile(r'(\d\d\d\d\d\d\d\d\d\d)').search(number) and re.match(r'0',number):
        ws['D{}'.format(i)]='Helpline_No'
        ws['D{}'.format(i)].font=Font(color="00FFFF")
    elif re.compile(r'(\d\d\d)-(\d\d\d\d\d\d\d\d)').search(number):
        ws['D{}'.format(i)]='Landline_No.'
        ws['D{}'.format(i)].font=Font(color="008000")
    elif re.compile(r'(\d\d\d\d)-(\d\d\d\d\d\d\d)').search(number):
        ws['D{}'.format(i)]='Landline_No.'
        ws['D{}'.format(i)].font=Font(color="008000")
    elif re.compile(r'(\d\d\d\d\d\d\d\d\d\d)').search(number):
        ws['D{}'.format(i)]='Mobile_No.'
        ws['D{}'.format(i)].font=Font(color="FF0000")
    elif re.compile(r'(\d\d\d)').search(number):
        ws['D{}'.format(i)]='Emergency_Ambulance_No.'
        ws['D{}'.format(i)].font=Font(color="0000FF")
    else:
        ws['D{}'.format(i)]='None'

wb.save(r"D:\py-assignments-new\py-assignments\assignment-2\pdfreader.xlsx")


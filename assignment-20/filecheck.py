from openpyxl.descriptors.excel import Percentage
import pandas as pd
import numpy as np
import tabula
import openpyxl

import sys
from tabula.io import read_pdf

from pdfminer.high_level import extract_text

from Levenshtein import distance, ratio


wb=openpyxl.load_workbook(r"D:\py-assignments-new\py-assignments\assignment-20\Phase2_SAP_DocMaterialNames (1).xlsx")
sheets=wb.sheetnames
# print(wb.active.title)


sh1=wb['Sheet2']  
sh1['C1'].value='Percentage(%)'
        
row=sh1.max_row
column=sh1.max_column




sh1['C1'].value='Percentage(%)'
for i in range(2,row+1):
    str1=sh1.cell(row=i,column=1).value
    if str1:
       data1=str1.upper()
    else:
        data1=" "
    
    for j in range(i,i+1):
        str2=sh1.cell(row=j,column=2).value
        if str2:
            data2=str2.upper()
        else:
            data2=" "
        
        a=ratio(data1,data2)*100
        # print(data1,"  ",data2,round(a,2))
        
        sh1['C{}'.format(i)]=round(a,2)

print("*****************")

str1=""
str2=""
sh1=wb['GRNDetails']
sh1['F1'].value='Percentage(%)'
for i in range(2,row+1):
    str1=sh1.cell(row=i,column=4).value
    if str1:
       data1=str1.upper()
    else:
        data1=" "
    
    for j in range(i,i+1):
        str2=sh1.cell(row=j,column=5).value
        if str2:
            data2=str2.upper()
        else:
            data2=" "
        
        a=ratio(data1,data2)*100
        # print(data1,"  ",data2,round(a,2))
        
        sh1['F{}'.format(i)]=round(a,2)
wb.save(r"D:\py-assignments-new\py-assignments\assignment-20\perc.xlsx")



# for i in range(2,row):
#     names=sh1.cell(row=i,column=4).value
#     for j in range(i,i+1):
#         name=sh1.cell(row=j,column=5).value
#         a=ratio(names,name)*100
#         print(name,"   ", names,"   ", round(a,2))

# print("*****************\n")

# for i in range(2,row):
#     names=sh1.cell(row=i,column=5).value
#     for j in range(i,i+1):
#         name=sh1.cell(row=j,column=4).value
#         a=ratio(name,names)*100
#         print(name,"   ", names,"   ", round(a,2))








#     print(sh1['Di'].value) #or print(sh1.cell(2,4).value)
# data1=sh1['D3'].value.upper()
# data2=sh1['E3'].value.upper()
# # print(data2.upper())
# a=ratio(data1,data2)*100
# print(data1, "in---", data2, round(a,2))

# a1=ratio(data2,data1)*100
# print(data2, "in--- ", data1, round(a1,2))



             

# file=r"D:\py-assignments-new\py-assignments\assignment-20\GRN Invoice files\download (1).pdf"
# data=tabula.io.read_pdf(file,pages=1)

with open(r'D:\py-assignments-new\py-assignments\assignment-20\download (1).txt') as f:
    text = f.readlines()
# for i in text:
#     print(i)

data="HYDROCHLORIC123"

for i in text:
    a=ratio(data,i)*100
    if a>50:
        print(i)
        print(round(a,2))
        

# wb=openpyxl.load_workbook(r"C:\Users\rdharavath\Downloads\GOOD RECEIPT -01062021 to 28.12.2021 (1).xlsx")
# sheets=wb.sheetnames
# print(wb.active.title)




wb=openpyxl.load_workbook(r"D:\py-assignments-new\py-assignments\assignment-20\Phase2_SAP_DocMaterialNames (1).xlsx")
sheets=wb.sheetnames
print(wb.active.title)

row=sh1.max_row
column=sh1.max_column

sh1=wb['GRNDetails']

data1="FORMIC ACID BULK"

percent=int(input("enter percentage : "))
for i in range(1,row+1):
    for j in range(1,column+1):
        str1=sh1.cell(row=i,column=j).value
        
        str2=str(str1)
        a=ratio(str2,data1)*100
        if a>percent:
            print(str2)
            print(round(a,2))
    # if str1:
    #    data1=str1.upper()
    # else:
    #     data1=" "
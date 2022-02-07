import openpyxl
import sys


wb=openpyxl.load_workbook(r"D:\app\Xref mapping data BASF Ver 2.0 (1).xlsx")
sheets=wb.sheetnames
# print(wb.active.title)
sh=wb['ConFig']
row=sh.max_row
column=sh.max_column

#input values
SAP_GROUP='WT005809'

SAP_Group_Counter=2

SAP_Characteristic=[]

for i in range(6,row+1):
    if SAP_GROUP==str(sh['F{}'.format(i)].value):
        SAP_Char = sh['H{}'.format(i)].value
        SAP_Characteristic.append(SAP_Char)
        
print(SAP_Characteristic)

# row=sh.max_row
# column=sh.max_column


wb=openpyxl.load_workbook(r"D:\app\PO4533737920_GR5000965446_BN88399847G0.xlsx")
sheets=wb.sheetnames
# print(wb.active.title)


sh1=wb['SAP_ILData']
sh1['B7'].value=SAP_GROUP
sh1['B8'].value=SAP_Group_Counter

sh2=wb['SAPChar']

i = 2
while i < len(SAP_Characteristic)+1:
    for j in range(i-2,len(SAP_Characteristic)):
        sh2.cell(row=i,column=1).value=SAP_Characteristic[j]
        print(sh2.cell(row=i,column=1).value)
        
        i = i + 1
        wb.save(r"D:\app\PO4533737920_GR5000965446_BN88399847G0.xlsx")
        wb.close()

# for i in SAP_Characteristic(2,len(SAP_Characteristic)+1):
    
    
#     print(i)
    # break
        # print(data[0:30])
        
        # sh2.cell(row=i,column=1).value = data
        # print(sh2.cell(row=i,column=1).value)
        # print(sh2['A{}'.format(i)].value )
    # print(i,data)
        
        # data1= data
        # print(data1)
        # break
    # print(j)
    # print(sh2.cell(row=i,column=1).value)  
    # sh2.cell(row=i,column=1).value=data1


# for i in range(1,row+1):
#     sh1['A{}'.format(i)]=SAP_Characteristic[0:i]
        

    
    

